import os
import datetime
import openpyxl
import mutagen
from mutagen.easyid3 import EasyID3
from mutagen.mp4 import MP4
# from mutagen import ID3NoHeaderError

from logger import setup_logger

logger = setup_logger(__name__)


class AudioSyncData:
    # 「設定」シートの行（0始まり）
    SYNC_DIR_FROM = 0
    EXCLUDE_DIR = 1
    INCLUDE_EXT = 2
    SYNC_DIR_TO = 3

    # 「Albums」「Not In Albums」、Playlistの列（1始まり）
    MSG = 1
    SYNC = 2
    TITLE = 3
    ARTIST = 4
    ALBUM_ARTIST = 5
    COMPOSER = 6
    ALBUM = 7
    TRACK_NUM = 8
    LENGTH = 9
    FILENAME = 10
    FILEPATH_FROM = 11
    FILEPATH_TO_RELATIVE = 12
    CODEC = 13
    ADDED_DATE = 15
    UPDATE_DATE = 14

    class Audio:
        # AudioSyncData内の「Albums」「Not in Albums」などに書かれている一つ一つの曲を表す
        # または、ディレクトリ内の実音楽ファイルを表す
        # 実ファイルを表すときは、引数を付けてnewする
        now = datetime.datetime.now()

        def __init__(self, filepath=None, relative_filepath=None):
            (
                self.msg,
                self.sync,
                self.title,
                self.artist,
                self.album_artist,
                self.composer,
                self.album,
                self.track_num,
                self.length,
                self.filename,
                self.filepath_from,
                self.filepath_to_relative,
                self.codec,
                self.added_date,
                self.update_date,
            ) = [None for _ in range(15)]

            if filepath is None or relative_filepath is None:
                return

            def if_key_error(eid3, tag):
                try:
                    return eid3[tag][0]
                except KeyError:
                    return ""

            self.__init__()
            self.filename = os.path.splitext(os.path.basename(filepath))[0]
            self.filepath_from = filepath
            self.filepath_to_relative = relative_filepath
            self.update_date = datetime.datetime.fromtimestamp(
                os.stat(filepath).st_mtime
            ).isoformat(sep=" ", timespec="seconds")
            self.added_date = self.now.isoformat(sep=" ", timespec="seconds")
            self.msg = None
            self.sync = None
            if os.path.splitext(filepath)[1] == ".mp3":
                self.codec = "mp3"  # TODO : 仮
                try:
                    eid3 = EasyID3(filepath)
                except mutagen.id3.ID3NoHeaderError:
                    self.msg = "!"
                    return

                self.title = if_key_error(eid3, "title")
                self.album = if_key_error(eid3, "album")
                self.artist = if_key_error(eid3, "artist")
                self.album_artist = if_key_error(eid3, "albumartist")
                self.composer = if_key_error(eid3, "composer")
                self.track_num = if_key_error(eid3, "tracknumber")
                self.length = if_key_error(eid3, "length")
                # TODO : 残件実装
                # rt. = if_key_error(eid3, "")
                # rt. = if_key_error(eid3, "")

            elif (
                os.path.splitext(filepath)[1] == ".mp4"
                or os.path.splitext(filepath)[1] == ".m4a"
            ):
                self.codec = "mp4"  # TODO : 仮
                mp4 = MP4(filepath)
                # TODO : 残件実装
                self.title = if_key_error(mp4.tags, "\xa9nam")
                self.album = if_key_error(mp4.tags, "\xa9alb")
                self.artist = if_key_error(mp4.tags, "\xa9ART")
                self.album_artist = if_key_error(mp4.tags, "aART")
                self.composer = if_key_error(mp4.tags, "\xa9wrt")
                self.track_num = if_key_error(mp4.tags, "trkn")
                if self.track_num != "":
                    if self.track_num[1] == 0:
                        self.track_num = self.track_num[0]
                    else:
                        self.track_num = (
                            str(min(self.track_num)) + "/" + str(max(self.track_num))
                        )
                self.length = mp4.info.length

            # TODO : shift-jisのタグに対しては「.encode("raw_unicode_escape").decode("shift_jis")」でutf8に戻せる
            # debug

    #            import chardet
    #            if chardet.detect(self.title.encode("raw_unicode_escape"))["encoding"] == "SHIFT_JIS" :
    #                print(self.title.encode("raw_unicode_escape").decode("shift_jis"))
    #            if chardet.detect(self.artist.encode("raw_unicode_escape"))["encoding"] == "SHIFT_JIS" :
    #                print(self.artist.encode("raw_unicode_escape").decode("shift_jis"))
    #            if chardet.detect(self.album.encode("raw_unicode_escape"))["encoding"] == "SHIFT_JIS" :
    #                print(self.album.encode("raw_unicode_escape").decode("shift_jis"))
    #            if chardet.detect(self.album_artist.encode("raw_unicode_escape"))["encoding"] == "SHIFT_JIS" :
    #                print(self.album_artist.encode("raw_unicode_escape").decode("shift_jis"))

    CACHE_SHEET_NAME = "_Cache"

    def __init__(self, filepath):
        # AudioSyncData.xlsxを表す
        #
        # filepath : 文字列。AudioSyncDataの場所（フルパス）
        self.__workbook = openpyxl.load_workbook(filepath)

        self.__sheet_Albums = None
        self.__sheet_Not_in_Albums = None
        self.__sheet_Cache = None

        # 隠しシート（キャッシュ）の初期化
        if self.CACHE_SHEET_NAME not in self.__workbook.sheetnames:
            self.__workbook.create_sheet(self.CACHE_SHEET_NAME)
            self.__workbook[self.CACHE_SHEET_NAME].sheet_state = "hidden"

            # ヘッダーをAlbumsシートからコピー
            src_sheet = self.__workbook["Albums"]
            dst_sheet = self.__workbook[self.CACHE_SHEET_NAME]
            for i in range(1, src_sheet.max_column + 1):
                dst_sheet.cell(row=1, column=i).value = src_sheet.cell(
                    row=1, column=i
                ).value

        self.__sheet_Cache = AudioSyncData.__Sheet(
            self.__workbook[self.CACHE_SHEET_NAME]
        )

    def __del__(self):
        pass

    @property
    def include_dir(self):
        # 「設定」シートに書いてある「対象ディレクトリ」
        #
        # 戻り値 : 文字列のイテレータ

        for item in self.__workbook["設定"].iter_cols(min_col=2):
            if item[self.SYNC_DIR_FROM].value is None:
                continue
            yield item[self.SYNC_DIR_FROM].value

    @property
    def include_extention(self):
        # 「設定」シートに書いてある「対象拡張子」
        #
        # 戻り値 : 文字列のイテレータ
        for item in self.__workbook["設定"].iter_cols(min_col=2):
            if item[self.INCLUDE_EXT].value is None:
                break
            yield "." + item[self.INCLUDE_EXT].value

    @property
    def dir_to_synchronize(self):
        # 「設定」シートに書いてある「同期先ディレクトリ」
        #
        # 戻り値 : 文字列のリスト
        rt = []
        for item in self.__workbook["設定"].iter_cols(min_col=2):
            if item[self.SYNC_DIR_TO].value is None:
                break
            rt.append(item[self.SYNC_DIR_TO].value)
        return rt

    def save(self, filename):
        self.__workbook.save(filename)

    class __Sheet:
        # 「Albums」「Not in Albums」などのシートを表す
        # シート内検索(インデックスでの参照)はファイルパスで行い、Audioインスタンスが返る
        # イテレーションではAudioインスタンスが返る

        def __init__(self, sheet, playlist=False):
            self.__sheet = sheet
            self.__is_playlist = playlist

            self.audio_dict = {}
            # {filepath : (num_row, Audio)}

            self.delete_empty_row()

            min_row = 4 if self.__is_playlist else 2
            for row in range(min_row, self.__sheet.max_row + 1):
                if (
                    self.__sheet.cell(row=row, column=self.header["ファイルパス"]).value
                    is None
                ):
                    continue
                audio = AudioSyncData.Audio()
                audio.msg = None
                audio.sync = self.__sheet.cell(
                    row=row, column=self.header["sync"]
                ).value
                audio.title = self.__sheet.cell(
                    row=row, column=self.header["タイトル"]
                ).value
                audio.artist = self.__sheet.cell(
                    row=row, column=self.header["アーティスト"]
                ).value
                audio.album_artist = self.__sheet.cell(
                    row=row, column=self.header["アルバムアーティスト"]
                ).value
                audio.composer = self.__sheet.cell(
                    row=row, column=self.header["作曲者"]
                ).value
                audio.album = self.__sheet.cell(
                    row=row, column=self.header["アルバム"]
                ).value
                audio.track_num = self.__sheet.cell(
                    row=row, column=self.header["#"]
                ).value
                audio.length = self.__sheet.cell(
                    row=row, column=self.header["長さ"]
                ).value
                audio.filename = self.__sheet.cell(
                    row=row, column=self.header["ファイル名"]
                ).value
                audio.filepath_from = self.__sheet.cell(
                    row=row, column=self.header["ファイルパス"]
                ).value
                audio.filepath_to_relative = self.__sheet.cell(
                    row=row, column=self.header["同期先相対ファイルパス"]
                ).value
                audio.codec = self.__sheet.cell(
                    row=row, column=self.header["コーデック"]
                ).value
                audio.update_date = self.__sheet.cell(
                    row=row, column=self.header["更新日時"]
                ).value
                audio.added_date = self.__sheet.cell(
                    row=row, column=self.header["追加日時"]
                ).value
                # audio. = self.__sheet.cell(row=row,column=AudioSyncData.).value
                self.audio_dict[
                    self.__sheet.cell(row=row, column=self.header["ファイルパス"]).value
                ] = (row, audio)
                logger.debug("read item from sheet. filepath=" + audio.filepath_from)

        @property
        def header(self):
            # シートの見出しを返す
            # 列番号
            # self.header[項目名] で 列番号 が取れる

            ###項目名一覧
            # msg
            # sync
            # タイトル
            # アーティスト
            # アルバムアーティスト
            # 作曲者
            # アルバム
            ##      <-これはトラック番号
            # 長さ
            # ファイル名
            # ファイルパス
            # 同期先相対ファイルパス
            # コーデック
            # 更新日時
            # 追加日時
            ###

            if not hasattr(self, "_header"):
                self._header = {}
                if not self.__is_playlist:
                    for i in range(1, self.__sheet.max_column + 1):
                        legend = self.__sheet.cell(1, i).value
                        self._header[legend] = i
                else:
                    # playlistは3行目が見出し
                    for i in range(1, self.__sheet.max_column + 1):
                        legend = self.__sheet.cell(3, i).value
                        self._header[legend] = i
            return self._header

        def delete_empty_row(self):
            for i in reversed(
                range(4 if self.__is_playlist else 2, self.__sheet.max_row)
            ):
                if (
                    self.__sheet.cell(row=i, column=self.header["ファイルパス"]).value
                    is None
                ):
                    self.__sheet.delete_rows(i)

        def __getitem__(self, filepath):
            _, rt = self.audio_dict[filepath]
            return rt

        def __setitem__(self, filepath, audio):
            # filepathの行が存在しているときはaudioの情報で上書き
            # 存在しない場合は行を新規追加

            if filepath in self.audio_dict:
                row, _ = self.audio_dict[filepath]
            else:
                row = self.__sheet.max_row + 1
            self.__sheet.cell(row=row, column=self.header["msg"]).value = audio.msg
            self.__sheet.cell(row=row, column=self.header["sync"]).value = audio.sync
            self.__sheet.cell(
                row=row, column=self.header["タイトル"]
            ).value = audio.title
            self.__sheet.cell(
                row=row, column=self.header["アーティスト"]
            ).value = audio.artist
            self.__sheet.cell(
                row=row, column=self.header["アルバムアーティスト"]
            ).value = audio.album_artist
            self.__sheet.cell(
                row=row, column=self.header["作曲者"]
            ).value = audio.composer
            self.__sheet.cell(
                row=row, column=self.header["アルバム"]
            ).value = audio.album
            self.__sheet.cell(row=row, column=self.header["#"]).value = audio.track_num
            self.__sheet.cell(row=row, column=self.header["長さ"]).value = audio.length
            self.__sheet.cell(
                row=row, column=self.header["ファイル名"]
            ).value = audio.filename
            self.__sheet.cell(
                row=row, column=self.header["ファイルパス"]
            ).value = audio.filepath_from
            self.__sheet.cell(
                row=row, column=self.header["同期先相対ファイルパス"]
            ).value = audio.filepath_to_relative
            self.__sheet.cell(
                row=row, column=self.header["コーデック"]
            ).value = audio.codec
            self.__sheet.cell(
                row=row, column=self.header["更新日時"]
            ).value = audio.update_date
            self.__sheet.cell(
                row=row, column=self.header["追加日時"]
            ).value = audio.added_date
            self.audio_dict[filepath] = (row, audio)
            logger.info("Data updated at row=" + str(row) + ". filepath=" + filepath)

        def __contains__(self, filepath):
            return filepath in self.audio_dict

        def __iter__(self):
            return (x for _, x in self.audio_dict.values())

        def make_m3u8(self, sep="\\"):
            # シートに書いてあるファイル類をm3u8プレイリストの文字列にして返す
            #
            # 戻り値 : 文字列。m3u8プレイリスト
            # TODO : 転送先においてあるファイルだけを書き出すようにしたい -> 本当に必要か？
            #       とりあえず、sync=○となっているものだけリスト化
            if self.__is_playlist == False:
                return ""
            rt = ""
            rt += "#EXTM3U\n\n"
            min_row = 4
            for row in range(min_row, self.__sheet.max_row + 1):
                try:
                    filepath = (
                        self.__sheet.cell(
                            row=row, column=self.header["同期先相対ファイルパス"]
                        )
                        .value[1:]
                        .replace(os.sep, sep)
                    )
                    title = self.__sheet.cell(
                        row=row, column=self.header["タイトル"]
                    ).value
                    sync = self.__sheet.cell(row=row, column=self.header["sync"]).value
                    if filepath != "" and sync == "○":
                        rt += "#EXTINF:-1," + title + "\n"
                        rt += filepath + "\n\n"
                except TypeError:
                    logger.debug(
                        'Playlist "'
                        + str(self.__sheet.cell(row=1, column=3).value)
                        + '" : processing skip of row '
                        + str(row)
                    )
            return rt

    @property
    def sheet_Albums(self):
        # 「Albums」シートに書いてあるデータ
        #
        # 戻り値 : __Sheetインスタンス
        if self.__sheet_Albums is None:
            self.__sheet_Albums = AudioSyncData.__Sheet(self.__workbook["Albums"])
        return self.__sheet_Albums

    @property
    def sheet_Not_in_Albums(self):
        # 「Not in Albums」シートに書いてあるデータ
        #
        # 戻り値 : __Sheetインスタンス
        if self.__sheet_Not_in_Albums is None:
            self.__sheet_Not_in_Albums = AudioSyncData.__Sheet(
                self.__workbook["Not in Albums"]
            )
        return self.__sheet_Not_in_Albums

    @property
    def sheets_playlist(self):
        # {プレイリスト名:Sheetインスタンス}
        rt = {}
        for ws in self.__workbook.worksheets:
            if (
                ws.cell(row=1, column=3).value == "プレイリスト名："
                and ws.cell(row=1, column=4).value is not None
                and ws.cell(row=2, column=4).value is not None
            ):
                rt[ws.cell(row=1, column=4).value] = AudioSyncData.__Sheet(
                    ws, playlist=True
                )
        return rt

    def get_audio_filepath_list(self):
        # 「対象ディレクトリ」に含まれる、拡張子が「対象拡張子」のファイル一覧をリストにして返す
        #
        # 戻り値 : tuple(String:ファイルのフルパス, String:同期先相対ファイルパス)のイテレータ。ファイル一覧（フルパス）

        def __get_audio_filepath_list(dir_path):
            for entry in os.scandir(dir_path):
                if entry.is_dir():
                    for rt in __get_audio_filepath_list(entry.path):
                        yield rt
                else:
                    if os.path.splitext(entry.path)[1] in self.include_extention:
                        yield entry.path

        for dir in self.include_dir:
            for rt in __get_audio_filepath_list(dir):
                yield (rt, rt[len(os.path.dirname(dir)) :])

    def get_audio_file_list(self):
        # get_audio_filepath_listの各要素をAudioインスタンスにして返す

        #
        # 戻り値 : Audioインスタンスのイテレータ

        for filepath, relative_filepath in self.get_audio_filepath_list():
            logger.debug("make Audio instatnce. filepath=" + filepath)
            rt = AudioSyncData.Audio(
                filepath=filepath, relative_filepath=relative_filepath
            )

            yield rt

    def update(self, update_all=False):
        # AudioSyncData.xlsx内の一覧を最新化する。
        # 一覧に存在しないファイルは追加する。
        # ファイルが存在しない場合はmsgに値を入れる
        # update_allがTrueのときは、すでに一覧に書かれているファイルでもタグ内容を最新化する

        # TODO : プレイリストもちゃんとupdateする
        albums_dict = {}
        not_in_albums_dict = {}
        # dictionary{ファイルパス : [Audioインスタンス, ファイル存在flag]}

        # シートのデータを読み込み
        for audio in self.sheet_Albums:
            albums_dict[audio.filepath_from] = [audio, False]
        for audio in self.sheet_Not_in_Albums:
            not_in_albums_dict[audio.filepath_from] = [audio, False]

        # キャッシュのデータを読み込み
        cache_dict = {}
        for audio in self.__sheet_Cache:
            cache_dict[audio.filepath_from] = audio

        n = 0
        for audio_filepath, relative_path in self.get_audio_filepath_list():
            n += 1

            # ディスク上のファイルの更新日時を取得
            current_mtime = datetime.datetime.fromtimestamp(
                os.stat(audio_filepath).st_mtime
            ).isoformat(sep=" ", timespec="seconds")

            # キャッシュにあるか、更新日時が新しいか確認
            use_cache = False
            if not update_all and audio_filepath in cache_dict:
                cached_audio = cache_dict[audio_filepath]
                if cached_audio.update_date == current_mtime:
                    use_cache = True

            if use_cache:
                # キャッシュを使用
                audio = cache_dict[audio_filepath]
                logger.debug("Use cache for: " + audio.filename)
            else:
                # ディスクから読み込み
                audio = AudioSyncData.Audio(
                    filepath=audio_filepath, relative_filepath=relative_path
                )
                logger.debug("Read from disk: " + audio.filename)
                # キャッシュを更新
                self.__sheet_Cache[audio.filepath_from] = audio

            # 表示用シートの更新・修復ロジック
            target_sheet = None
            target_dict = None

            if audio.album is None or audio.album == "":
                target_sheet = self.sheet_Not_in_Albums
                target_dict = not_in_albums_dict
            else:
                target_sheet = self.sheet_Albums
                target_dict = albums_dict

            if audio.filepath_from in target_dict:
                # すでに一覧にある場合
                target_dict[audio.filepath_from][1] = True
                existing_audio = target_dict[audio.filepath_from][0]

                # ユーザー編集可能な列（sync, added_dateなど）は維持する
                # ただし、メタデータ（タイトル、アーティストなど）はキャッシュ（＝正しい値）で上書きして修復する

                # syncとadded_dateを既存のものからコピー
                audio.sync = existing_audio.sync
                audio.added_date = existing_audio.added_date

                # もし既存のデータと新しいデータ（キャッシュ由来）が異なれば、修復とみなして更新
                # Audioクラスの比較メソッドがないので、主要な属性で比較するか、無条件で上書きするか。
                # ここでは無条件で上書きすることで「修復」を実現する。
                # ただし、無駄な書き込みを減らすために値の比較はしたほうが良いが、
                # Audioオブジェクトの属性が多いので、一旦は常に上書きする（openpyxl側で値が同じなら変更されないことを期待、あるいは__setitem__でログが出る）

                target_sheet[audio.filepath_from] = audio
            else:
                # 一覧にない場合は追加
                # added_dateを現在時刻に更新（表示シートから削除されていた場合の再追加にも対応）
                audio.added_date = AudioSyncData.Audio.now.isoformat(
                    sep=" ", timespec="seconds"
                )
                print("add audio file : " + audio.filename)
                target_sheet[audio.filepath_from] = audio

        # ファイルが存在しない場合は、msgに"-"を記入
        # TODO : とりあえずの対応
        for filepath, (audio, flag) in albums_dict.items():
            if flag is False:
                audio = self.sheet_Albums[filepath]
                audio.msg = "-"
                self.sheet_Albums[filepath] = audio
        for filepath, (audio, flag) in not_in_albums_dict.items():
            if flag is False:
                audio = self.sheet_Not_in_Albums[filepath]
                audio.msg = "-"
                self.sheet_Not_in_Albums[filepath] = audio
