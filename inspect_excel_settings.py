import openpyxl


def inspect_settings(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "設定" not in wb.sheetnames:
        print("Sheet '設定' not found.")
        return

    ws = wb["設定"]
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")

    for row in range(1, ws.max_row + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            row_values.append(str(val))
        print(f"Row {row}: {', '.join(row_values)}")


if __name__ == "__main__":
    inspect_settings("src/audiosync/AudioSyncData.xlsx")
